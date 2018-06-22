import openpyxl
import shutil
from os.path import exists
import re
columns=shutil.get_terminal_size().columns
def signup():
  b=openpyxl.load_workbook('fin.xlsx')
  acc=b['account']
  pas=b['passbook']
  maxr=acc.max_row
  flag=0
  n=[]
  n.append(input("Enter your First Name: "))
  n.append(input("Enter your Last Name: "))
  n.append(input("Enter your Username: "))
  i=2
  while i<maxr+1:
    if n[2]==acc.cell(row=i,column=3).value:
      print("Username already exists")
      n[2]=input("Enter new Username: ")
      i=1
    i+=1
  n.append(input("Enter your email: "))
  while(flag==0):
    p=re.compile('.+@.+\..+')
    m=p.match(n[3])
    if m: 
      flag=1
      i=2
      while i<maxr+1:
        if n[3]==acc.cell(row=i,column=4).value:
          print("Email already exists")
          n[3]=input("Enter new Email: ")
          i=2
          flag=0
          break
        i+=1
    else:
      n[3]=input("Enter a valid formatted Email: ")
  n.append(input("Enter Password: "))
  q=re.compile('[a-z]')
  r=re.compile('[A-Z]') 
  s=re.compile('[0-9]')
  t=re.compile('\W')
  flag=0
  while(flag==0):
    m=q.search(n[4])
    p=r.search(n[4])
    o=s.search(n[4])
    u=t.search(n[4])
    if m and p and o and u and len(n[4])>7:
      flag=1
    if flag!=1:
      n[4]=input("Enter strong password: ")
  for i in range(0,5):
    acc.cell(row=maxr+1,column=i+1).value=n[i]
  f=pas.max_row
  pas.cell(row=f+1,column=1).value=n[2]
  pas.cell(row=f+1,column=3).value=pas.cell(row=f+1,column=4).value=10000
  b.save("fin.xlsx")
  print("Account Created Successfully!!")
def login():
  b=openpyxl.load_workbook('fin.xlsx')
  acc=b['account']
  pas=b['passbook']
  maxr=acc.max_row
  flag=0
  fl=0
  j=2#j points to that row in which username and password exists in account sheet...if exists
  i=input("Enter Username: ")
  p=input("Enter Password: ")
  while j<maxr+1:
        if i==acc.cell(row=j,column=3).value and p==acc.cell(row=j,column=5).value:
          fl=1
          break
        j+=1
  if(fl==0): print("Invalid username or password")
  while fl==1:
    print(("HELLO "+acc.cell(row=j,column=1).value).center(columns))
    print("1. Debit\n2. Credit\n3. Change Password\n4. Transfer\n5. Logout")
    c=int(input())
    if c==5:
      fl=0
      return
    flag=0
    k=pas.max_row#k points to that row in which username and final details exists in balance sheet...if exists(k wil exist if program is at this line..
    f=k#f points to max_row in balance sheet
    while k>1 and flag==0:
      if pas.cell(row=k,column=1).value==i:
        flag=1
        break
      k-=1
    if c==1:
      a=int(input("Enter amount: "))
      pas.cell(row=f+1,column=1).value=pas.cell(row=k,column=1).value
      pas.cell(row=f+1,column=2).value=a
      pas.cell(row=f+1,column=4).value=pas.cell(row=k,column=4).value-a
      print("Done!!")
    elif c==2:
      a=int(input("Enter amount: "))
      pas.cell(row=f+1,column=1).value=pas.cell(row=k,column=1).value
      pas.cell(row=f+1,column=3).value=a
      pas.cell(row=f+1,column=4).value=pas.cell(row=k,column=4).value+a
      print("Done!!")
    elif c==3:
      a=input("Enter new Password: ")
      qa=re.compile('[a-z]')
      ra=re.compile('[A-Z]') 
      sa=re.compile('[0-9]')
      ta=re.compile('\W')
      flag=0
      while(flag==0):
        ma=qa.search(a)
        pa=ra.search(a)
        oa=sa.search(a)
        ua=ta.search(a)
        if ma and pa and oa and ua and len(a)>7:
          flag=1
        if flag!=1:
          a=input("Enter strong password: ")###
      acc.cell(row=j,column=5).value=a
      print("Done!!")
    elif c==4:
      a=int(input("Enter amount: "))
      n=input("Enter Username(to whom): ")
      flag=0
      l=f#l points to user to whom amount will be transfered
      while l>1 and flag==0:
        if pas.cell(row=l,column=1).value==n:
          flag=1
          pas.cell(row=f+1,column=1).value=pas.cell(row=k,column=1).value
          pas.cell(row=f+1,column=2).value=a
          pas.cell(row=f+1,column=4).value=pas.cell(row=k,column=4).value-a
          f=f+1
          pas.cell(row=f+1,column=1).value=pas.cell(row=l,column=1).value
          pas.cell(row=f+1,column=3).value=a
          pas.cell(row=f+1,column=4).value=pas.cell(row=l,column=4).value+a
          print("Done!!")
          break
        l-=1
      if flag==0: print("Invalid username!!")   
    b.save("fin.xlsx")
if not exists('fin.xlsx'):
  b=openpyxl.Workbook()
  ac1=b.create_sheet('account')
  ba1=b.create_sheet('passbook')
  shee=b.sheetnames
  b.remove(b[shee[0]])
  col=['A','B','C','D','E']
  for col1 in col:
    ac1.column_dimensions[col1].width=30
    ba1.column_dimensions[col1].width=30
  ac1.cell(row=1,column=1).value='First Name'
  ac1.cell(row=1,column=2).value='Last Name'
  ac1.cell(row=1,column=3).value='Username'
  ac1.cell(row=1,column=4).value='Email'
  ac1.cell(row=1,column=5).value='Password'
  ba1.cell(row=1,column=1).value='Username'
  ba1.cell(row=1,column=2).value='Debit'
  ba1.cell(row=1,column=3).value='Credit'
  ba1.cell(row=1,column=4).value='Balance'
  b.save("fin.xlsx")
while 1:
  print("NET BANKING".center(columns))#print at center of terminal
  print("1. SIGN UP\n2. LOGIN\n3. Exit")
  choice=int(input())
  if(choice==1):
    print("signup")
    signup()
  elif(choice==2):
    login()
  elif(choice==3):
    break
